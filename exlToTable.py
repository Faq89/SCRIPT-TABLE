import argparse
import os
import openpyxl
import pandas as pd


def excel_to_html_table(excel_file_path, sheet_name=None, num_columns=None):
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    last_row = worksheet.max_row
    for row in reversed(range(1, worksheet.max_row + 1)):
        if any(worksheet.cell(row=row, column=col).value is not None for col in range(1, worksheet.max_column + 1)):
            last_row = row
            break

    merged_cells = worksheet.merged_cells.ranges
    merged_dict = {}
    for merged_cell in merged_cells:
        min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_dict[(row, col)] = (rowspan, colspan) if (row, col) == (min_row, min_col) else (None, None)

    html = '<table class="table-list">\n'
    html += '  <tr>\n'

    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1) if worksheet.cell(row=1, column=col).value is not None]
    if num_columns:
        headers = headers[:num_columns]

    for header in headers:
        html += f'    <th>{header}</th>\n'
    html += '  </tr>\n'

    for row in worksheet.iter_rows(min_row=2, max_row=last_row, max_col=len(headers)):
        html += '  <tr>\n'
        for cell in row:
            cell_value = cell.value if cell.value is not None else ''
            if isinstance(cell_value, float):
                cell_value = f"{int(cell_value)}" if cell_value.is_integer() else f"{cell_value:.2f}"
            else:
                cell_value = str(cell_value)

            if (cell.row, cell.column) in merged_dict:
                rowspan, colspan = merged_dict[(cell.row, cell.column)]
                if rowspan and colspan:
                    rowspan_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                    colspan_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                    style_attr = ' style="vertical-align: middle;"'
                    html += f'    <td{rowspan_attr}{colspan_attr}{style_attr}>{cell_value}</td>\n'
            else:
                html += f'    <td>{cell_value}</td>\n'
        html += '  </tr>\n'

    html += '</table>'
    return html


def generate_output_filename(sheet_name, directory):
    base_name = sheet_name if sheet_name else "Output"
    filename = f"{base_name}.html"
    file_path = os.path.join(directory, filename)
    counter = 1
    while os.path.exists(file_path):
        filename = f"{base_name}{counter}.html"
        file_path = os.path.join(directory, filename)
        counter += 1
    return file_path


def ask_to_replace_file(filename):
    while True:
        user_input = input(f"The file '{filename}' already exists. Do you want to replace it? (yes/no): ").strip().lower()
        if user_input == 'yes':
            return True
        elif user_input == 'no':
            return False


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert an Excel table to HTML.")
    parser.add_argument("excel_file", help="The path to the Excel file")
    parser.add_argument("--sheet", help="Excel sheet name (optional)", default=None)
    parser.add_argument("--num_columns", type=int, help="Number of columns to process (optional)", default=None)
    args = parser.parse_args()

    try:
        script_directory = os.path.dirname(os.path.abspath(__file__))
        if args.num_columns is None:
            workbook = openpyxl.load_workbook(args.excel_file)
            worksheet = workbook[args.sheet] if args.sheet else workbook.active
            args.num_columns = len([cell for cell in worksheet[1] if cell.value is not None])

        table_html = excel_to_html_table(args.excel_file, sheet_name=args.sheet, num_columns=args.num_columns)
        output_filename = generate_output_filename(args.sheet, script_directory)

        if os.path.exists(output_filename):
            if ask_to_replace_file(output_filename):
                with open(output_filename, 'w', encoding='utf-8') as f:
                    f.write(table_html)
                print(f"The HTML table has been saved in '{output_filename}'")
            else:
                output_filename = generate_output_filename(args.sheet, script_directory)
                with open(output_filename, 'w', encoding='utf-8') as f:
                    f.write(table_html)
                print(f"The HTML table has been saved in '{output_filename}'")
        else:
            with open(output_filename, 'w', encoding='utf-8') as f:
                f.write(table_html)
            print(f"The HTML table has been saved in '{output_filename}'")

    except Exception as e:
        print(f"Error: {e}")
